"""Bulk endpoint import (CSV/Excel) and export.

Import is a two-step flow on purpose: ``analyse`` validates a file and returns
a preview with per-row errors and duplicate detection, and ``commit`` writes
only the rows the operator confirmed. Nothing is written during analysis, so a
bad file cannot leave half an import behind.
"""

from __future__ import annotations

import csv
import io
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Iterable, Sequence

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging import get_logger
from app.models.endpoint import Endpoint
from app.monitoring.validators import (
    UrlValidationError,
    clamp_interval,
    clamp_timeout,
    normalise_status_codes,
    parse_target,
)
from app.services import endpoint_service

logger = get_logger(__name__)

CSV_TEMPLATE_COLUMNS = [
    "name",
    "url",
    "environment",
    "tags",
    "interval",
    "timeout",
    "description",
    "owner",
    "team",
    "application",
    "method",
    "expected_status",
    "check_type",
    "monitoring_enabled",
    "ssl_monitoring",
    "verify_ssl",
    "follow_redirects",
    "failure_threshold",
    "response_time_threshold_ms",
]

# Accept the header spellings people actually produce from spreadsheets.
_COLUMN_ALIASES: dict[str, str] = {
    "name": "name",
    "endpoint": "name",
    "endpoint_name": "name",
    "endpointname": "name",
    "url": "url",
    "endpoint_url": "url",
    "target": "url",
    "hostname": "url",
    "host": "url",
    "address": "url",
    "environment": "environment",
    "env": "environment",
    "tags": "tags",
    "tag": "tags",
    "labels": "tags",
    "interval": "interval",
    "interval_seconds": "interval",
    "monitoring_interval": "interval",
    "check_interval": "interval",
    "timeout": "timeout",
    "timeout_seconds": "timeout",
    "description": "description",
    "notes": "description",
    "owner": "owner",
    "owner_email": "owner",
    "team": "team",
    "application": "application",
    "app": "application",
    "method": "method",
    "http_method": "method",
    "expected_status": "expected_status",
    "expected_status_code": "expected_status",
    "expected_status_codes": "expected_status",
    "expected_http_status": "expected_status",
    "check_type": "check_type",
    "type": "check_type",
    "monitoring_enabled": "monitoring_enabled",
    "enabled": "monitoring_enabled",
    "active": "monitoring_enabled",
    "ssl_monitoring": "ssl_monitoring",
    "ssl_monitoring_enabled": "ssl_monitoring",
    "ssl": "ssl_monitoring",
    "verify_ssl": "verify_ssl",
    "ssl_verify": "verify_ssl",
    "follow_redirects": "follow_redirects",
    "redirects": "follow_redirects",
    "failure_threshold": "failure_threshold",
    "retries": "failure_threshold",
    "response_time_threshold_ms": "response_time_threshold_ms",
    "response_time_threshold": "response_time_threshold_ms",
    "latency_threshold_ms": "response_time_threshold_ms",
}

_TRUTHY = {"1", "true", "yes", "y", "on", "enabled", "enable"}
_FALSY = {"0", "false", "no", "n", "off", "disabled", "disable"}


class ImportError_(ValueError):
    """The file itself could not be read."""


@dataclass
class RowResult:
    row_number: int
    raw: dict[str, Any]
    payload: dict[str, Any] | None = None
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    duplicate_of: str | None = None

    @property
    def is_valid(self) -> bool:
        return not self.errors and self.payload is not None

    def as_dict(self) -> dict[str, Any]:
        return {
            "row_number": self.row_number,
            "name": (self.payload or {}).get("name") or self.raw.get("name"),
            "url": (self.payload or {}).get("url") or self.raw.get("url"),
            "environment": (self.payload or {}).get("environment"),
            "tags": (self.payload or {}).get("tags") or [],
            "interval_seconds": (self.payload or {}).get("interval_seconds"),
            "timeout_seconds": (self.payload or {}).get("timeout_seconds"),
            "valid": self.is_valid,
            "errors": self.errors,
            "warnings": self.warnings,
            "duplicate_of": self.duplicate_of,
        }


@dataclass
class ImportAnalysis:
    filename: str
    total_rows: int = 0
    rows: list[RowResult] = field(default_factory=list)
    file_errors: list[str] = field(default_factory=list)
    detected_columns: list[str] = field(default_factory=list)
    unknown_columns: list[str] = field(default_factory=list)

    @property
    def valid_rows(self) -> list[RowResult]:
        return [r for r in self.rows if r.is_valid]

    @property
    def invalid_rows(self) -> list[RowResult]:
        return [r for r in self.rows if not r.is_valid]

    def as_dict(self) -> dict[str, Any]:
        return {
            "filename": self.filename,
            "total_rows": self.total_rows,
            "valid_count": len(self.valid_rows),
            "invalid_count": len(self.invalid_rows),
            "duplicate_count": sum(1 for r in self.rows if r.duplicate_of),
            "detected_columns": self.detected_columns,
            "unknown_columns": self.unknown_columns,
            "file_errors": self.file_errors,
            "rows": [r.as_dict() for r in self.rows],
        }


# ------------------------------------------------------------ file reading
def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ImportError_("file encoding could not be detected; save it as UTF-8 CSV")


def read_csv(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = _decode_text(data)
    sample = text[:8192]
    try:
        dialect: Any = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # plain comma-separated

    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if reader.fieldnames is None:
        raise ImportError_("the file appears to be empty")
    headers = [(h or "").strip() for h in reader.fieldnames]
    rows = []
    for record in reader:
        rows.append(
            {
                (k or "").strip(): ("" if v is None else str(v))
                for k, v in record.items()
                if k is not None
            }
        )
    return headers, rows


def read_excel(data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError_(
            "Excel support requires the openpyxl package"
        ) from exc

    try:
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ImportError_(f"could not read the Excel file: {exc}") from exc

    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration as exc:
            raise ImportError_("the spreadsheet is empty") from exc

        headers = [
            str(cell).strip() if cell is not None else "" for cell in header_row
        ]
        rows: list[dict[str, str]] = []
        for values in iterator:
            if values is None or all(v is None or str(v).strip() == "" for v in values):
                continue
            record: dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = values[index] if index < len(values) else None
                if isinstance(value, datetime):
                    value = value.isoformat()
                record[header] = "" if value is None else str(value).strip()
            rows.append(record)
        return headers, rows
    finally:
        workbook.close()


def read_file(filename: str, data: bytes) -> tuple[list[str], list[dict[str, str]]]:
    lowered = (filename or "").lower()
    if lowered.endswith((".xlsx", ".xlsm")):
        return read_excel(data)
    if lowered.endswith(".xls"):
        raise ImportError_(
            "legacy .xls files are not supported - save the sheet as .xlsx or .csv"
        )
    if lowered.endswith((".csv", ".txt", ".tsv")) or not lowered:
        return read_csv(data)
    # Fall back on content sniffing: an .xlsx is a ZIP archive.
    if data[:2] == b"PK":
        return read_excel(data)
    return read_csv(data)


# ------------------------------------------------------------- row parsing
def _normalise_headers(headers: Sequence[str]) -> tuple[dict[str, str], list[str]]:
    """Map raw headers onto canonical field names."""
    mapping: dict[str, str] = {}
    unknown: list[str] = []
    for header in headers:
        if not header:
            continue
        key = header.strip().lower().replace(" ", "_").replace("-", "_")
        canonical = _COLUMN_ALIASES.get(key)
        if canonical:
            mapping[header] = canonical
        else:
            unknown.append(header)
    return mapping, unknown


def _to_bool(value: Any, default: bool | None = None) -> bool | None:
    if value is None:
        return default
    text = str(value).strip().lower()
    if not text:
        return default
    if text in _TRUTHY:
        return True
    if text in _FALSY:
        return False
    raise ValueError(f"'{value}' is not a yes/no value")


def _to_int(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        raise ValueError(f"'{value}' is not a number") from None


def _split_tags(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = [
        piece.strip().lower()
        for chunk in text.replace(";", ",").replace("|", ",").split(",")
        for piece in [chunk]
        if piece.strip()
    ]
    seen: set[str] = set()
    ordered: list[str] = []
    for part in parts:
        if part not in seen:
            seen.add(part)
            ordered.append(part)
    return ordered


def parse_row(
    row_number: int,
    record: dict[str, Any],
    mapping: dict[str, str],
    *,
    config: dict[str, Any],
) -> RowResult:
    """Validate one spreadsheet row into an endpoint payload."""
    canonical: dict[str, Any] = {}
    for header, value in record.items():
        field_name = mapping.get(header)
        if field_name and str(value).strip() != "":
            canonical[field_name] = str(value).strip()

    result = RowResult(row_number=row_number, raw=canonical)

    name = canonical.get("name", "").strip()
    url_value = canonical.get("url", "").strip()

    if not name and not url_value:
        result.errors.append("row is empty")
        return result
    if not url_value:
        result.errors.append("url is required")
    if not name:
        # A missing name is recoverable: derive one from the host so the
        # operator is not blocked on a cosmetic field.
        if url_value:
            try:
                name = parse_target(url_value).hostname
                result.warnings.append(f"name was empty; derived '{name}' from the URL")
            except UrlValidationError:
                result.errors.append("name is required")
        else:
            result.errors.append("name is required")
    if len(name) > 160:
        result.errors.append("name must be at most 160 characters")

    target = None
    if url_value:
        try:
            target = parse_target(url_value)
        except UrlValidationError as exc:
            result.errors.append(str(exc))

    interval = None
    timeout = None
    try:
        raw_interval = _to_int(canonical.get("interval"))
        interval = clamp_interval(raw_interval or config.get("default_monitor_interval"))
        if raw_interval and raw_interval != interval:
            result.warnings.append(
                f"interval {raw_interval}s adjusted to {interval}s by the "
                "configured limits"
            )
    except ValueError as exc:
        result.errors.append(f"interval: {exc}")

    try:
        raw_timeout = _to_int(canonical.get("timeout"))
        timeout = clamp_timeout(
            raw_timeout or config.get("default_timeout"), interval=interval
        )
        if raw_timeout and raw_timeout != timeout:
            result.warnings.append(
                f"timeout {raw_timeout}s adjusted to {timeout}s"
            )
    except ValueError as exc:
        result.errors.append(f"timeout: {exc}")

    expected_status = None
    if canonical.get("expected_status"):
        try:
            expected_status = normalise_status_codes(canonical["expected_status"])
        except UrlValidationError as exc:
            result.errors.append(f"expected_status: {exc}")

    booleans: dict[str, bool | None] = {}
    for field_name, default in (
        ("monitoring_enabled", True),
        ("ssl_monitoring", True),
        ("verify_ssl", True),
        ("follow_redirects", True),
    ):
        try:
            booleans[field_name] = _to_bool(canonical.get(field_name), default)
        except ValueError as exc:
            result.errors.append(f"{field_name}: {exc}")
            booleans[field_name] = default

    numbers: dict[str, int | None] = {}
    for field_name in ("failure_threshold", "response_time_threshold_ms"):
        try:
            numbers[field_name] = _to_int(canonical.get(field_name))
        except ValueError as exc:
            result.errors.append(f"{field_name}: {exc}")
            numbers[field_name] = None

    method = (canonical.get("method") or "GET").upper()
    if method not in {"GET", "HEAD", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"}:
        result.errors.append(f"method '{method}' is not supported")

    check_type = (canonical.get("check_type") or "http").lower()
    if check_type not in {"http", "tcp", "tls"}:
        result.errors.append(f"check_type '{check_type}' is not supported")

    tags = _split_tags(canonical.get("tags"))
    if len(tags) > 20:
        result.warnings.append("only the first 20 tags were kept")
        tags = tags[:20]

    if result.errors:
        return result

    result.payload = {
        "name": name,
        "url": target.url if target else url_value,
        "environment": canonical.get("environment"),
        "create_missing_environment": True,
        "tags": tags,
        "interval_seconds": interval,
        "timeout_seconds": timeout,
        "description": canonical.get("description"),
        "owner": canonical.get("owner"),
        "team": canonical.get("team"),
        "application": canonical.get("application"),
        "http_method": method,
        "check_type": check_type,
        "expected_status_codes": expected_status,
        "monitoring_enabled": booleans.get("monitoring_enabled", True),
        "ssl_monitoring_enabled": booleans.get("ssl_monitoring", True),
        "verify_ssl": booleans.get("verify_ssl", True),
        "follow_redirects": booleans.get("follow_redirects", True),
        "failure_threshold": numbers.get("failure_threshold"),
        "response_time_threshold_ms": numbers.get("response_time_threshold_ms"),
    }
    return result


# --------------------------------------------------------------- analysis
async def analyse(
    session: AsyncSession,
    *,
    filename: str,
    data: bytes,
    config: dict[str, Any],
    max_rows: int,
) -> ImportAnalysis:
    """Validate an uploaded file and build a preview. Writes nothing."""
    analysis = ImportAnalysis(filename=filename or "upload")

    try:
        headers, records = read_file(filename, data)
    except ImportError_ as exc:
        analysis.file_errors.append(str(exc))
        return analysis

    mapping, unknown = _normalise_headers(headers)
    analysis.detected_columns = sorted(set(mapping.values()))
    analysis.unknown_columns = unknown

    if "url" not in mapping.values():
        analysis.file_errors.append(
            "no 'url' column was found. Expected headers include: "
            + ", ".join(CSV_TEMPLATE_COLUMNS[:6])
        )
        return analysis

    analysis.total_rows = len(records)
    if len(records) > max_rows:
        analysis.file_errors.append(
            f"file contains {len(records)} rows; the limit is {max_rows}. "
            "Split the file and import it in batches."
        )
        return analysis
    if not records:
        analysis.file_errors.append("the file contains no data rows")
        return analysis

    # Row 1 is the header, so data starts at row 2 - matching what the
    # operator sees in their spreadsheet.
    for index, record in enumerate(records, start=2):
        analysis.rows.append(parse_row(index, record, mapping, config=config))

    await _detect_duplicates(session, analysis)
    return analysis


async def _detect_duplicates(
    session: AsyncSession, analysis: ImportAnalysis
) -> None:
    """Flag rows that clash with the database or with each other."""
    candidates = [r for r in analysis.rows if r.payload]
    if not candidates:
        return

    urls = {r.payload["url"].lower() for r in candidates}
    names = {r.payload["name"].strip().lower() for r in candidates}

    existing_rows = (
        await session.execute(
            select(Endpoint.name, Endpoint.url).where(
                func.lower(Endpoint.url).in_(urls)
                | func.lower(Endpoint.name).in_(names)
            )
        )
    ).all()
    existing_urls = {row[1].lower() for row in existing_rows}
    existing_names = {row[0].strip().lower() for row in existing_rows}

    seen_urls: dict[str, int] = {}
    seen_names: dict[str, int] = {}

    for row in candidates:
        url = row.payload["url"].lower()
        name = row.payload["name"].strip().lower()

        if url in existing_urls:
            row.errors.append("an endpoint with this URL already exists")
            row.duplicate_of = "database"
            continue
        if name in existing_names:
            row.errors.append("an endpoint with this name already exists")
            row.duplicate_of = "database"
            continue
        if url in seen_urls:
            row.errors.append(f"duplicate URL, also on row {seen_urls[url]}")
            row.duplicate_of = f"row {seen_urls[url]}"
            continue
        if name in seen_names:
            row.errors.append(f"duplicate name, also on row {seen_names[name]}")
            row.duplicate_of = f"row {seen_names[name]}"
            continue
        seen_urls[url] = row.row_number
        seen_names[name] = row.row_number


# ----------------------------------------------------------------- commit
@dataclass
class ImportOutcome:
    created: list[dict[str, Any]] = field(default_factory=list)
    failed: list[dict[str, Any]] = field(default_factory=list)
    skipped: list[dict[str, Any]] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        return {
            "created_count": len(self.created),
            "failed_count": len(self.failed),
            "skipped_count": len(self.skipped),
            "created": self.created,
            "failed": self.failed,
            "skipped": self.skipped,
        }


async def commit(
    session: AsyncSession,
    analysis: ImportAnalysis,
    *,
    config: dict[str, Any],
    created_by_id: uuid.UUID | None,
    only_rows: Iterable[int] | None = None,
) -> ImportOutcome:
    """Create endpoints for the valid rows.

    Each row is created inside its own SAVEPOINT so one late failure (a race
    against a concurrent create, say) does not discard the rows that already
    succeeded.
    """
    outcome = ImportOutcome()
    wanted = set(only_rows) if only_rows is not None else None

    for row in analysis.rows:
        if wanted is not None and row.row_number not in wanted:
            continue
        if not row.is_valid:
            outcome.skipped.append(
                {
                    "row_number": row.row_number,
                    "name": row.raw.get("name"),
                    "url": row.raw.get("url"),
                    "errors": row.errors,
                }
            )
            continue

        try:
            async with session.begin_nested():
                endpoint = await endpoint_service.create_endpoint(
                    session,
                    row.payload,
                    config=config,
                    created_by_id=created_by_id,
                )
            outcome.created.append(
                {
                    "row_number": row.row_number,
                    "id": str(endpoint.id),
                    "name": endpoint.name,
                    "url": endpoint.url,
                }
            )
        except Exception as exc:
            outcome.failed.append(
                {
                    "row_number": row.row_number,
                    "name": row.payload.get("name"),
                    "url": row.payload.get("url"),
                    "error": str(exc),
                }
            )
            logger.warning(
                "import_row_failed", row=row.row_number, error=str(exc)
            )

    logger.info(
        "import_completed",
        created=len(outcome.created),
        failed=len(outcome.failed),
        skipped=len(outcome.skipped),
    )
    return outcome


# ----------------------------------------------------------------- export
EXPORT_COLUMNS = [
    "name",
    "url",
    "protocol",
    "hostname",
    "port",
    "check_type",
    "method",
    "environment",
    "tags",
    "description",
    "owner",
    "team",
    "application",
    "monitoring_enabled",
    "paused",
    "interval",
    "timeout",
    "expected_status",
    "follow_redirects",
    "verify_ssl",
    "ssl_monitoring",
    "failure_threshold",
    "response_time_threshold_ms",
    "auth_type",
    "current_status",
    "last_status_code",
    "last_response_time_ms",
    "last_checked_at",
    "ssl_status",
    "ssl_issuer",
    "ssl_expires_at",
    "ssl_days_remaining",
    "created_at",
    "updated_at",
]


def _export_row(endpoint: Endpoint) -> list[Any]:
    """One export row.

    Credentials are represented only by their type - never the secret, not even
    the masked hint, because an export leaves the application.
    """
    return [
        endpoint.name,
        endpoint.url,
        endpoint.protocol,
        endpoint.hostname,
        endpoint.port,
        endpoint.check_type,
        endpoint.http_method,
        endpoint.environment.name if endpoint.environment else "",
        ",".join(endpoint.tag_names),
        endpoint.description or "",
        endpoint.owner or "",
        endpoint.team or "",
        endpoint.application or "",
        "yes" if endpoint.monitoring_enabled else "no",
        "yes" if endpoint.is_paused else "no",
        endpoint.interval_seconds,
        endpoint.timeout_seconds,
        endpoint.expected_status_codes,
        "yes" if endpoint.follow_redirects else "no",
        "yes" if endpoint.verify_ssl else "no",
        "yes" if endpoint.ssl_monitoring_enabled else "no",
        endpoint.failure_threshold,
        endpoint.response_time_threshold_ms or "",
        endpoint.auth_type,
        endpoint.current_status,
        endpoint.last_status_code or "",
        round(endpoint.last_response_time_ms, 2)
        if endpoint.last_response_time_ms is not None
        else "",
        _iso(endpoint.last_checked_at),
        endpoint.ssl_status,
        endpoint.ssl_issuer or "",
        _iso(endpoint.ssl_expires_at),
        endpoint.ssl_days_remaining if endpoint.ssl_days_remaining is not None else "",
        _iso(endpoint.created_at),
        _iso(endpoint.updated_at),
    ]


def _iso(value: datetime | None) -> str:
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.isoformat()


def export_csv(endpoints: Sequence[Endpoint]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(EXPORT_COLUMNS)
    for endpoint in endpoints:
        writer.writerow(_export_row(endpoint))
    # A BOM keeps Excel from mangling non-ASCII names on open.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def export_excel(endpoints: Sequence[Endpoint]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Endpoints"

    sheet.append(EXPORT_COLUMNS)
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"

    for endpoint in endpoints:
        sheet.append(_export_row(endpoint))

    for index, column in enumerate(EXPORT_COLUMNS, start=1):
        width = max(
            len(column) + 2,
            min(
                48,
                max(
                    (
                        len(str(sheet.cell(row=r, column=index).value or ""))
                        for r in range(2, min(sheet.max_row, 200) + 1)
                    ),
                    default=12,
                )
                + 2,
            ),
        )
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(EXPORT_COLUMNS))}{max(sheet.max_row, 1)}"
    )

    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def csv_template() -> bytes:
    """A ready-to-fill template with one example row."""
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, lineterminator="\r\n")
    writer.writerow(CSV_TEMPLATE_COLUMNS)
    writer.writerow(
        [
            "Translation API",
            "https://api.example.com/health",
            "production",
            "backend,critical",
            "60",
            "10",
            "Translation backend health probe",
            "platform@example.com",
            "Platform",
            "Bhashini",
            "GET",
            "200",
            "http",
            "yes",
            "yes",
            "yes",
            "yes",
            "3",
            "2000",
        ]
    )
    writer.writerow(
        [
            "Portal",
            "https://portal.example.com",
            "production",
            "frontend",
            "60",
            "10",
            "Main portal",
            "web@example.com",
            "Web",
            "Portal",
            "GET",
            "200,301",
            "http",
            "yes",
            "yes",
            "yes",
            "yes",
            "3",
            "",
        ]
    )
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


async def recheck_duplicates(
    session: AsyncSession, analysis: ImportAnalysis
) -> None:
    """Public wrapper: re-run duplicate detection before committing.

    Endpoints can be created by someone else between the preview and the
    confirmation, so the check is repeated rather than trusted from the
    stored preview.
    """
    await _detect_duplicates(session, analysis)
