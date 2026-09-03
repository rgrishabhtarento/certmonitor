"""Bulk CSV/Excel import and configuration export."""

from __future__ import annotations

import csv
import io

from sqlalchemy import func, select

from app.models.endpoint import Endpoint

VALID_CSV = b"""name,url,environment,tags,interval,timeout,description
Translation API,https://api.example.com/health,production,"backend,critical",60,10,Translation backend
Portal,https://portal.example.com,production,"frontend",60,10,Main portal
Dev API,https://dev-api.example.com/health,development,"backend,dev",120,10,Development API
"""


async def upload(client, headers, data: bytes, filename="endpoints.csv"):
    return await client.post(
        "/api/import",
        headers=headers,
        files={"file": (filename, data, "text/csv")},
    )


class TestPreview:
    async def test_valid_file_previews_every_row(self, client, admin_headers):
        response = await upload(client, admin_headers, VALID_CSV)
        assert response.status_code == 200, response.text

        body = response.json()
        assert body["total_rows"] == 3
        assert body["valid_count"] == 3
        assert body["invalid_count"] == 0
        assert body["token"]
        assert len(body["rows"]) == 3
        assert body["rows"][0]["row_number"] == 2  # row 1 is the header
        assert body["rows"][0]["tags"] == ["backend", "critical"]

    async def test_preview_writes_nothing(self, client, admin_headers, session):
        await upload(client, admin_headers, VALID_CSV)
        count = (await session.execute(select(func.count(Endpoint.id)))).scalar()
        assert count == 0

    async def test_missing_url_column_is_rejected(self, client, admin_headers):
        response = await upload(
            client, admin_headers, b"name,environment\nOnly a name,production\n"
        )
        assert response.status_code == 200
        assert response.json()["file_errors"]
        assert "url" in response.json()["file_errors"][0]

    async def test_empty_file_is_rejected(self, client, admin_headers):
        response = await upload(client, admin_headers, b"")
        assert response.status_code == 400

    async def test_unsupported_extension_is_rejected(self, client, admin_headers):
        response = await upload(client, admin_headers, b"x", filename="config.pdf")
        assert response.status_code == 415

    async def test_invalid_rows_are_flagged_individually(self, client, admin_headers):
        data = (
            b"name,url\n"
            b"Good,https://good.example.com/health\n"
            b"Bad scheme,ftp://bad.example.com\n"
            b"Bad host,https://-nope-.example.com\n"
        )
        response = await upload(client, admin_headers, data)
        body = response.json()

        assert body["valid_count"] == 1
        assert body["invalid_count"] == 2
        invalid = [row for row in body["rows"] if not row["valid"]]
        assert all(row["errors"] for row in invalid)

    async def test_missing_name_is_derived_with_a_warning(
        self, client, admin_headers
    ):
        """A cosmetic gap should not block an import."""
        response = await upload(
            client, admin_headers, b"name,url\n,https://derived.example.com/health\n"
        )
        row = response.json()["rows"][0]

        assert row["valid"] is True
        assert row["name"] == "derived.example.com"
        assert any("derived" in warning for warning in row["warnings"])

    async def test_duplicate_within_the_file_is_flagged(self, client, admin_headers):
        data = (
            b"name,url\n"
            b"First,https://same.example.com/health\n"
            b"Second,https://same.example.com/health\n"
        )
        response = await upload(client, admin_headers, data)
        body = response.json()

        assert body["valid_count"] == 1
        assert body["duplicate_count"] == 1
        second = body["rows"][1]
        assert second["valid"] is False
        assert "row 2" in second["duplicate_of"]

    async def test_duplicate_against_the_database_is_flagged(
        self, client, admin_headers
    ):
        await client.post(
            "/api/endpoints",
            json={"name": "Existing", "url": "https://existing.example.com/health"},
            headers=admin_headers,
        )

        response = await upload(
            client,
            admin_headers,
            b"name,url\nNew name,https://existing.example.com/health\n",
        )
        row = response.json()["rows"][0]
        assert row["valid"] is False
        assert row["duplicate_of"] == "database"

    async def test_alternative_column_names_are_recognised(
        self, client, admin_headers
    ):
        data = (
            b"endpoint_name,target,env,labels,interval_seconds\n"
            b"Aliased,https://aliased.example.com/health,staging,api,300\n"
        )
        response = await upload(client, admin_headers, data)
        row = response.json()["rows"][0]

        assert row["valid"] is True
        assert row["name"] == "Aliased"
        assert row["interval_seconds"] == 300
        assert row["tags"] == ["api"]

    async def test_unknown_columns_are_reported_not_fatal(
        self, client, admin_headers
    ):
        data = b"name,url,cost_centre\nX,https://x.example.com/h,CC-1\n"
        response = await upload(client, admin_headers, data)
        body = response.json()

        assert body["valid_count"] == 1
        assert "cost_centre" in body["unknown_columns"]

    async def test_aggressive_interval_is_adjusted_with_a_warning(
        self, client, admin_headers
    ):
        data = b"name,url,interval\nFast,https://fast.example.com/h,1\n"
        response = await upload(client, admin_headers, data)
        row = response.json()["rows"][0]

        assert row["valid"] is True
        assert row["interval_seconds"] == 30
        assert any("adjusted" in warning for warning in row["warnings"])

    async def test_semicolon_delimited_file_is_detected(self, client, admin_headers):
        data = b"name;url;environment\nSemi;https://semi.example.com/h;production\n"
        response = await upload(client, admin_headers, data)
        assert response.json()["valid_count"] == 1

    async def test_bom_prefixed_utf8_is_handled(self, client, admin_headers):
        data = "﻿name,url\nBOM,https://bom.example.com/h\n".encode("utf-8")
        response = await upload(client, admin_headers, data)
        assert response.json()["valid_count"] == 1

    async def test_viewer_cannot_import(self, client, viewer_headers):
        response = await upload(client, viewer_headers, VALID_CSV)
        assert response.status_code == 403


class TestConfirm:
    async def test_confirm_creates_the_endpoints(
        self, client, admin_headers, session
    ):
        preview = await upload(client, admin_headers, VALID_CSV)
        token = preview.json()["token"]

        response = await client.post(
            "/api/import/confirm", json={"token": token}, headers=admin_headers
        )
        assert response.status_code == 200, response.text

        body = response.json()
        assert body["created_count"] == 3
        assert body["failed_count"] == 0

        count = (await session.execute(select(func.count(Endpoint.id)))).scalar()
        assert count == 3

    async def test_imported_endpoints_carry_their_configuration(
        self, client, admin_headers
    ):
        preview = await upload(client, admin_headers, VALID_CSV)
        await client.post(
            "/api/import/confirm",
            json={"token": preview.json()["token"]},
            headers=admin_headers,
        )

        listing = await client.get(
            "/api/endpoints", params={"search": "Translation"}, headers=admin_headers
        )
        endpoint = listing.json()["items"][0]
        assert endpoint["environment"]["name"] == "production"
        assert sorted(tag["name"] for tag in endpoint["tags"]) == [
            "backend",
            "critical",
        ]
        assert endpoint["interval_seconds"] == 60

    async def test_a_subset_can_be_imported(self, client, admin_headers, session):
        preview = await upload(client, admin_headers, VALID_CSV)
        body = preview.json()
        chosen = [body["rows"][0]["row_number"]]

        response = await client.post(
            "/api/import/confirm",
            json={"token": body["token"], "row_numbers": chosen},
            headers=admin_headers,
        )
        assert response.json()["created_count"] == 1

        count = (await session.execute(select(func.count(Endpoint.id)))).scalar()
        assert count == 1

    async def test_invalid_rows_are_skipped_not_failed(self, client, admin_headers):
        data = (
            b"name,url\n"
            b"Good,https://good.example.com/health\n"
            b"Bad,ftp://bad.example.com\n"
        )
        preview = await upload(client, admin_headers, data)

        response = await client.post(
            "/api/import/confirm",
            json={"token": preview.json()["token"]},
            headers=admin_headers,
        )
        body = response.json()
        assert body["created_count"] == 1
        assert body["skipped_count"] == 1

    async def test_an_unknown_token_is_gone(self, client, admin_headers):
        response = await client.post(
            "/api/import/confirm",
            json={"token": "no-such-token"},
            headers=admin_headers,
        )
        assert response.status_code == 410

    async def test_a_token_cannot_be_reused(self, client, admin_headers):
        preview = await upload(client, admin_headers, VALID_CSV)
        token = preview.json()["token"]

        first = await client.post(
            "/api/import/confirm", json={"token": token}, headers=admin_headers
        )
        assert first.status_code == 200

        second = await client.post(
            "/api/import/confirm", json={"token": token}, headers=admin_headers
        )
        assert second.status_code == 410

    async def test_reimporting_the_same_file_creates_nothing(
        self, client, admin_headers, session
    ):
        """Re-running an import must be a no-op, not a duplicate fleet."""
        first_preview = await upload(client, admin_headers, VALID_CSV)
        await client.post(
            "/api/import/confirm",
            json={"token": first_preview.json()["token"]},
            headers=admin_headers,
        )

        second_preview = await upload(client, admin_headers, VALID_CSV)
        body = second_preview.json()
        assert body["valid_count"] == 0
        assert body["duplicate_count"] == 3

        count = (await session.execute(select(func.count(Endpoint.id)))).scalar()
        assert count == 3

    async def test_import_is_audited(self, client, admin_headers):
        preview = await upload(client, admin_headers, VALID_CSV)
        await client.post(
            "/api/import/confirm",
            json={"token": preview.json()["token"]},
            headers=admin_headers,
        )

        audit = await client.get(
            "/api/audit-logs",
            params={"action": "endpoints_imported"},
            headers=admin_headers,
        )
        assert audit.json()["meta"]["total"] == 1


class TestTemplate:
    async def test_template_is_downloadable_and_reimportable(
        self, client, admin_headers
    ):
        response = await client.get("/api/import/template", headers=admin_headers)
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]

        # The template must itself be a valid import file.
        preview = await upload(client, admin_headers, response.content, "template.csv")
        assert preview.json()["valid_count"] == 2


class TestExport:
    async def test_csv_export_contains_the_endpoint(self, client, admin_headers):
        await client.post(
            "/api/endpoints",
            json={
                "name": "Exported",
                "url": "https://exported.example.com/health",
                "tags": ["backend"],
            },
            headers=admin_headers,
        )

        response = await client.get(
            "/api/export", params={"format": "csv"}, headers=admin_headers
        )
        assert response.status_code == 200
        assert "attachment" in response.headers["content-disposition"]

        text = response.content.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
        assert len(rows) == 1
        assert rows[0]["name"] == "Exported"
        assert rows[0]["url"] == "https://exported.example.com/health"
        assert rows[0]["tags"] == "backend"

    async def test_export_never_contains_a_credential(self, client, admin_headers):
        """An export leaves the application, so no secret may be in it."""
        await client.post(
            "/api/endpoints",
            json={
                "name": "With secret",
                "url": "https://secret.example.com/health",
                "auth_type": "bearer",
                "auth_secret": "do-not-export-me",
            },
            headers=admin_headers,
        )

        response = await client.get(
            "/api/export", params={"format": "csv"}, headers=admin_headers
        )
        text = response.content.decode("utf-8-sig")

        assert "do-not-export-me" not in text
        # Not even the masked hint.
        assert "****" not in text
        # The type is exported so the configuration remains understandable.
        assert "bearer" in text

    async def test_excel_export_is_a_valid_workbook(self, client, admin_headers):
        await client.post(
            "/api/endpoints",
            json={"name": "Excel", "url": "https://excel.example.com/health"},
            headers=admin_headers,
        )

        response = await client.get(
            "/api/export", params={"format": "xlsx"}, headers=admin_headers
        )
        assert response.status_code == 200
        # xlsx files are ZIP archives.
        assert response.content[:2] == b"PK"

        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        assert sheet.max_row == 2
        assert sheet.cell(row=2, column=1).value == "Excel"
        workbook.close()

    async def test_export_honours_filters(self, client, admin_headers):
        for name, host in (("Keep", "keep"), ("Drop", "drop")):
            await client.post(
                "/api/endpoints",
                json={"name": name, "url": f"https://{host}.example.com/h"},
                headers=admin_headers,
            )

        response = await client.get(
            "/api/export",
            params={"format": "csv", "search": "Keep"},
            headers=admin_headers,
        )
        text = response.content.decode("utf-8-sig")
        assert "Keep" in text
        assert "Drop" not in text

    async def test_a_viewer_may_export(self, client, viewer_headers):
        """Read-only users still need to take a configuration snapshot."""
        response = await client.get(
            "/api/export", params={"format": "csv"}, headers=viewer_headers
        )
        assert response.status_code == 200

    async def test_export_is_audited(self, client, admin_headers):
        await client.get(
            "/api/export", params={"format": "csv"}, headers=admin_headers
        )
        audit = await client.get(
            "/api/audit-logs",
            params={"action": "endpoints_exported"},
            headers=admin_headers,
        )
        assert audit.json()["meta"]["total"] == 1
